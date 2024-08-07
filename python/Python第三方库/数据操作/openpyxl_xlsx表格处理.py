"""pygame  模块"""
import openpyxl
import openpyxl.styles
import openpyxl.utils


'''
1.xlsx表格处理

导入方式:
    import openpyxl
'''

print('1.')
print('——————————————————————————————————————————————————')

print('1.')

print('——————————————————————————————————————————————————')
print('\n')


'''
2.load_workbook(), Workbook()

load_workbook(filename, read_only=False, keep_vba=KEEP_VBA,
              data_only=False, keep_links=True): 加载表格
Workbook(write_only=False, iso_dates=False): 创建表格
以上实例的属性:
    []: 键入sheet名字获取sheet工作表
    worksheets: 包含所有sheet工作表的列表，通过下标来获取sheet
    sheetnames: 包含所有sheet工作表的名字的列表
    create_sheet(title=None, index=None): 创建sheet
    move_sheet(sheet, offset=0): 移动sheet，形参sheet为sheet或sheet的名字
    copy_worksheet(from_worksheet): 复制sheet
    save(filename): 保存到指定文件
    close(): 关闭，只影响只读或只写模式
    >>> del wb[]: 移除sheet工作表
'''

print('2.')
print('——————————————————————————————————————————————————')

# wb = openpyxl.load_workbook("表格.xlsx")
wb = openpyxl.Workbook()
print('1.', "实例:", wb)
print('2.', "worksheets:", wb.worksheets)
wb.create_sheet("sheet2")
print("3.", "create_sheet('sheet2'):", wb.worksheets)
del wb["sheet2"]
print("4.", "del wb['sheet2']:", wb.worksheets)


print('——————————————————————————————————————————————————')
print('\n')


'''
3.表格中的工作表sheet

声明:
    wb == openpyxl.Workbook() or load_workbook()
    sheet == wb.worksheets[0]
sheet:
    []: 下标格式: "A1", "A2", ...，获取单元格
    title: sheet的名字
    rows: 所有行，以列行排列
    columns: 所有列，以行列排列
    max_row: 最大行
    max_column: 最大列
    values: 所有单元格的值，以行列排列
    cell(row, column, value=None): 单元格
    >>> column_dimensions[].width: 宽
    >>> row_dimensions[].height: 高
'''

print('3.')
print('——————————————————————————————————————————————————')

sheet = wb.worksheets[0]
print("1.", sheet)
print('2.', "cell(row, col, 'value')")
sheet.cell(1, 1, "A1 第一行第一列")
sheet.cell(1, 2, "B1 第一行第二列")
sheet.cell(2, 1, "A2 第二行第一列")

print("3.", "全部行，第一层数组以列排列，第二层数组以行排列")
print([i for i in sheet.rows])

print("4.", "全部列，第一层数组以行排列，第二层数组以列排列")
print([i for i in sheet.columns])

print("5.", "所有单元格的值，第一层数组以行分隔，第二层数组以列分隔")
print([i for i in sheet.values])

print("6.", "最长行和最长列", (sheet.max_row, sheet.max_column))

print("7.", "设置单元格宽高")
sheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = len(sheet.cell(1, 1).value.encode("GBK"))
sheet.column_dimensions[openpyxl.utils.get_column_letter(2)].width = len(sheet.cell(1, 2).value.encode("GBK"))
sheet.row_dimensions[1].height = 24

print('——————————————————————————————————————————————————')
print('\n')


'''
4.sheet工作表中的cell单元格

声明:
    wb == openpyxl.Workbook() or load_workbook()
    sheet == wb.worksheets[0]
    cell = sheet.cell(1, 1) or sheet.cell["A1"]
cell:
    value: 值
    style: 格式，使用 openpyxl.styles 模块的 NamedStayle 实例 或者合法的 str 赋值
        合法的 str 有:
            ('Normal', 'Comma', 'Currency', 'Percent',
             'Comma [0]', 'Currency [0]', 'Hyperlink', 'Followed Hyperlink',
             'Note', 'Warning Text', 'Title',
             'Headline 1', 'Headline 2', 'Headline 3', 'Headline 4',
             'Input', 'Output', 'Calculation', 'Check Cell', 'Linked Cell',
             'Total', 'Good', 'Bad', 'Neutral',
             'Accent1', '20 % - Accent1', '40 % - Accent1', '60 % - Accent1',
             'Accent2', '20 % - Accent2', '40 % - Accent2', '60 % - Accent2',
             'Accent3', '20 % - Accent3', '40 % - Accent3', '60 % - Accent3',
             'Accent4', '20 % - Accent4', '40 % - Accent4', '60 % - Accent4',
             'Accent5', '20 % - Accent5', '40 % - Accent5', '60 % - Accent5',
             'Accent6', '20 % - Accent6', '40 % - Accent6', '60 % - Accent6',
             'Explanatory Text', 'Pandas')
    font: 字体，使用 openpyxl.styles 模块的 Font 实例赋值
    alignment: 对齐，使用 openpyxl.styles 模块的 Alignment 实例赋值
    fill: 颜色填充，使用 openpyxl.styles 模块的 PatternFill 实例赋值
    border: 边框，使用 openpyxl.styles 模块的 border 实例赋值
'''

print('4.')
print('——————————————————————————————————————————————————')

cell = sheet.cell(1, 1)
print('1.', cell)
print("2.", "cell.vlaue:", cell.value)
print("3.", "cell.style:", cell.style)
print("4.", "cell.font:", cell.font)
print("5.", "cell.alignment:", cell.alignment)
print("6.", "cell.fill:", cell.fill)
print("7.", "cell.border:", cell.border)

print('——————————————————————————————————————————————————')
print('\n')


'''
5.openpyxl.Styles

样式模块
属性:
    NamedStyle(name="Normal", font=Font(), fill=PatternFill(),
               border=Border(), alignment=Alignment(), number_format=None,
               protection=Protection(), builtinId=None, hidden=False,
               xfId=None):
        参数:
            name: 模式
            font: openpyxl.styles.Font
            fill: openpyxl.styles.PatternFill
            border: openpyxl.styles.Border
            alignment: openpyxl.styles.Alignment
    Font()
    Alignment()
    PatternFill(patternType=None, fgColor=Color(), bgColor=Color(),
                fill_type=None, start_color=None, end_color=None): 填充
        参数:
            patternType: 填充模式，可选值:
                ('lightGrid', 'lightVertical', 'darkGray', 'gray0625',
                 'lightHorizontal', 'darkDown', 'darkUp', 'lightUp',
                 'darkGrid', 'darkTrellis', 'lightDown', 'darkVertical',
                 'gray125', 'mediumGray', 'lightGray', 'darkHorizontal',
                 'lightTrellis', 'solid')
            fgColor: 背景颜色
            bgColor: ...
            fill_type: 同 patternType
            start_color: 同 fgColor
            end_color: 同 bgColor
    Border()
'''

print('5.')
print('——————————————————————————————————————————————————')

cell = sheet.cell(1, 1)
print(openpyxl.styles.NamedStyle)
print(openpyxl.styles.Font)
print(openpyxl.styles.Alignment)
print(openpyxl.styles.PatternFill)
print(openpyxl.styles.Border)

print('——————————————————————————————————————————————————')
print('\n')


wb.save("openpyxl.xlsx")
